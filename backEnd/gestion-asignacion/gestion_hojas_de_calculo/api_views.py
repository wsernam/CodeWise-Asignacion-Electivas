from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.parsers import MultiPartParser, FormParser
from tablib import Dataset
import openpyxl 
import io
import traceback 
import requests
import logging
import uuid
from django.core.cache import cache
import xlrd
from decimal import Decimal, InvalidOperation

# Importamos el nuevo Resource
from .resources import PerfilAcademicoResource 
from referencias.models import SeleccionEstudianteElectiva


logger = logging.getLogger(__name__)

class ValidarExcelAPIView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, format=None):
        excel_files = request.FILES.getlist('files')
        if not excel_files:
            return Response({'error': 'No se encontraron archivos en la solicitud.'}, status=status.HTTP_400_BAD_REQUEST)

        # --- INICIO: Guardado en Caché ---
        cache_key = f"excel_files_{uuid.uuid4()}"
        files_to_cache = []
        for f in excel_files:
            f.seek(0)
            files_to_cache.append({
                'name': f.name,
                'content': f.read(),
                'content_type': f.content_type
            })
        cache.set(cache_key, files_to_cache, timeout=3600) 
        # --- FIN: Guardado en Caché ---

        # 1. Obtener el periodo activo desde el otro microservicio
        try:
            periodo_response = requests.get('http://gestion-asignacion:8000/api/asignacion/procesos/periodo-activo/')
            if periodo_response.status_code == 204:
                return Response({'error': 'No se encontró un proceso de asignación ACTIVO.'}, status=status.HTTP_404_NOT_FOUND)
            periodo_response.raise_for_status()
            periodo_data = periodo_response.json()
            anio_activo = periodo_data['pa_anio']
            semestre_activo = periodo_data['pa_num_semestre']
        except requests.RequestException as e:
            return Response({'error': f'No se pudo comunicar con el servicio de asignación: {e}'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

        # 2. Obtener códigos de estudiante de la BD para el periodo activo
        codigos_db = set(SeleccionEstudianteElectiva.objects.filter(
            sel_anio=anio_activo,
            sel_num_semestre=semestre_activo
        ).values_list('est_codigo', flat=True).distinct())

        # 3. Extraer códigos de estudiante de los Excels USANDO EL RESOURCE
        codigos_excel = set()
        resource = PerfilAcademicoResource()
        errores_procesamiento = []

        # FUNCIÓN AUXILIAR: Detectar si una fila está completamente vacía
        def es_fila_vacia(row_dict):
            """
            Retorna True si todos los valores de la fila están vacíos.
            """
            if not row_dict:
                return True
            return all(
                val is None or str(val).strip() == '' 
                for val in row_dict.values()
            )

        for excel_file in excel_files:
            try:
                filename = excel_file.name.lower()
                if filename.endswith('.xlsx'):
                    file_format = 'xlsx'
                elif filename.endswith('.xls'):
                    file_format = 'xls'
                else:
                    errores_procesamiento.append(f"Archivo '{excel_file.name}' ignorado: formato no soportado.")
                    continue

                excel_file.seek(0)
                file_content = excel_file.read()

                dataset = Dataset()
                dataset.load(file_content, format=file_format)
                
                for i, row in enumerate(dataset.dict):
                    # SOLUCIÓN PROBLEMA 1: Saltar filas completamente vacías
                    if es_fila_vacia(row):
                        continue
                    
                    try:
                        estudiante_instance = resource.fields['est_codigo'].clean(row)
                        if estudiante_instance:
                            codigos_excel.add(estudiante_instance.est_codigo)
                    except ValueError as e:
                        # SOLUCIÓN PROBLEMA 2: Mensaje más específico
                        error_msg = str(e)
                        codigo_estudiante = row.get('CODIGO', 'desconocido')
                        
                        # Limpiar el código si viene en formato float
                        if codigo_estudiante != 'desconocido':
                            try:
                                str_code = str(codigo_estudiante)
                                if str_code.endswith('.0'):
                                    codigo_estudiante = str_code[:-2]
                                else:
                                    codigo_estudiante = str(int(float(str_code)))
                            except (ValueError, TypeError):
                                pass
                        
                        if 'no existe' in error_msg.lower() or 'not found' in error_msg.lower():
                            msg = f"Fila {i+2} del archivo '{excel_file.name}': El estudiante con código {codigo_estudiante} no está registrado en la base de datos."
                        else:
                            msg = f"Fila {i+2} del archivo '{excel_file.name}': Error de validación - {e}"
                        logger.warning(msg)
                        errores_procesamiento.append(msg)
                    except Exception as e:
                        # Capturar específicamente el error de Django "DoesNotExist"
                        error_msg = str(e)
                        codigo_estudiante = row.get('CODIGO', 'desconocido')
                        
                        # Limpiar el código si viene en formato float (ej: 123456.0 -> 123456)
                        if codigo_estudiante != 'desconocido':
                            try:
                                str_code = str(codigo_estudiante)
                                if str_code.endswith('.0'):
                                    codigo_estudiante = str_code[:-2]
                                else:
                                    codigo_estudiante = str(int(float(str_code)))
                            except (ValueError, TypeError):
                                pass
                        
                        if 'does not exist' in error_msg.lower() or 'matching query' in error_msg.lower():
                            msg = f"Fila {i+2} del archivo '{excel_file.name}': El estudiante con código {codigo_estudiante} no está registrado en la base de datos."
                        else:
                            msg = f"Fila {i+2} del archivo '{excel_file.name}': Error inesperado - {e}"
                        logger.warning(msg)
                        errores_procesamiento.append(msg)
                        
            except Exception as e:
                # SOLUCIÓN PROBLEMA 2: Mensaje más claro y amigable para el usuario
                msg = f"No se pudo procesar el archivo '{excel_file.name}'"
                
                # Detectar errores específicos comunes con mensajes amigables
                if 'xlrd' in str(e).lower():
                    msg += ". El archivo parece estar en formato XLS antiguo. Por favor, ábrelo en Excel y guárdalo como XLSX (Excel moderno)."
                elif 'openpyxl' in str(e).lower():
                    msg += ". Hubo un problema al leer el archivo XLSX. Verifica que no esté corrupto o protegido con contraseña."
                elif 'corrupt' in str(e).lower() or 'formato' in str(e).lower():
                    msg += ". El archivo parece estar dañado o en un formato no compatible. Intenta guardarlo nuevamente desde Excel."
                elif 'password' in str(e).lower() or 'encrypted' in str(e).lower():
                    msg += ". El archivo está protegido con contraseña. Por favor, remueve la protección antes de subirlo."
                else:
                    msg += f". Error: {str(e)[:100]}"
                    
                logger.error(f"Error procesando '{excel_file.name}': {e}", exc_info=True)
                errores_procesamiento.append(msg)

        # 4. Comparar y generar respuesta
        faltantes = list(codigos_db - codigos_excel)
        sobrantes = list(codigos_excel - codigos_db)
        coinciden = not faltantes and not sobrantes

        return Response({
            'cache_key': cache_key,
            'faltantes': faltantes,
            'sobrantes': sobrantes,
            'num_faltantes': len(faltantes),
            'num_sobrantes': len(sobrantes),
            'coinciden': coinciden,
            'periodo_evaluado': f'{anio_activo}-{semestre_activo}',
            'advertencias': errores_procesamiento
        }, status=status.HTTP_200_OK)
        
        
class CompletarYProcesarAPIView(APIView):
    def post(self, request, format=None):
        cache_key = request.data.get('cache_key')
        filas_a_completar = request.data.get('filas_a_completar', [])

        if not cache_key:
            return Response({'error': 'No se proporcionó un `cache_key`.'}, status=status.HTTP_400_BAD_REQUEST)

        # 1. Recuperar archivos de la caché
        cached_files = cache.get(cache_key)
        if not cached_files:
            return Response({'error': 'Los archivos han expirado o la clave es inválida. Por favor, vuelva a subirlos.'}, status=status.HTTP_404_NOT_FOUND)

        # 2. Identificar todas las filas incompletas originales
        vista_previsualizar = PrevisualizarIncompletosAPIView()
        # Simulamos una request para reutilizar la lógica de previsualización
        from django.core.files.uploadedfile import InMemoryUploadedFile
        
        temp_files_for_preview = []
        for file_data in cached_files:
            file_stream = io.BytesIO(file_data['content'])
            temp_file = InMemoryUploadedFile(file_stream, None, file_data['name'], file_data['content_type'], len(file_data['content']), None)
            temp_files_for_preview.append(temp_file)

        # CORRECCIÓN: El MockRequest debe simular la estructura de request.FILES,
        # que es un objeto con un método getlist, no un diccionario simple.
        class MockRequest:
            class MockFILES:
                def __init__(self, files):
                    self._files = files
                def getlist(self, key):
                    return self._files.get(key, [])
            FILES = MockFILES({'files': temp_files_for_preview})

        mock_request = MockRequest()
        preview_response = vista_previsualizar.post(mock_request)
        todas_las_filas_incompletas = preview_response.data.get('filas_incompletas', [])

        # 3. Determinar qué filas eliminar
        filas_completadas_set = {(item['archivo'], item['fila']) for item in filas_a_completar}
        filas_a_eliminar = [
            item for item in todas_las_filas_incompletas 
            if (item['archivo'], item['fila']) not in filas_completadas_set
        ]

        # 4. Editar los archivos Excel en memoria
        archivos_procesados_para_importar = []
        for file_data in cached_files:
            try:
                file_content_stream = io.BytesIO(file_data['content'])
                file_name = file_data['name']
                
                # --- INICIO: Lógica para manejar .xls y .xlsx ---
                # Si es un archivo .xls, lo convertimos a .xlsx en memoria antes de procesar.
                if file_name.lower().endswith('.xls'):
                    logger.info(f"Detectado archivo .xls ('{file_name}'). Convirtiendo a .xlsx en memoria.")
                    # Leer .xls con xlrd
                    xls_book = xlrd.open_workbook(file_contents=file_content_stream.read())
                    # Crear un nuevo workbook .xlsx con openpyxl
                    workbook = openpyxl.Workbook()
                    sheet = workbook.active
                    
                    # Copiar datos celda por celda
                    xls_sheet = xls_book.sheet_by_index(0)
                    for row in range(xls_sheet.nrows):
                        for col in range(xls_sheet.ncols):
                            sheet.cell(row=row + 1, column=col + 1).value = xls_sheet.cell_value(row, col)
                else:
                    # Si ya es .xlsx, simplemente lo cargamos
                    workbook = openpyxl.load_workbook(file_content_stream)
                sheet = workbook.active
                # --- FIN: Lógica para manejar .xls y .xlsx ---

                # Mapeo de cabeceras a índices de columna (1-based)
                headers = {cell.value: cell.column for cell in sheet[1]}

                # --- Eliminar filas (de abajo hacia arriba para no afectar índices) ---
                filas_para_este_archivo = sorted(
                    [item['fila'] for item in filas_a_eliminar if item['archivo'] == file_name],
                    reverse=True
                )
                for row_num in filas_para_este_archivo:
                    sheet.delete_rows(row_num)
                    logger.info(f"Fila {row_num} eliminada del archivo '{file_name}'.")

                # --- Actualizar filas ---
                filas_para_actualizar = [item for item in filas_a_completar if item['archivo'] == file_name]
                for item in filas_para_actualizar:
                    row_num = item['fila']
                    datos = item['datos']
                    for col_name, value in datos.items():
                        if col_name in headers:
                            col_idx = headers[col_name]
                            sheet.cell(row=row_num, column=col_idx, value=value)
                    logger.info(f"Fila {row_num} actualizada en el archivo '{file_name}'.")

                # Guardar el workbook modificado en un stream de bytes
                output_stream = io.BytesIO()
                workbook.save(output_stream)
                output_stream.seek(0)
                
                # Cambiamos la extensión a .xlsx si era un .xls para que la importación funcione
                final_file_name = file_name if not file_name.lower().endswith('.xls') else file_name[:-4] + '.xlsx'

                archivos_procesados_para_importar.append(
                    ('files', (final_file_name, output_stream.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))
                )

            except Exception as e:
                logger.error(f"Error procesando el archivo en memoria '{file_name}': {e}", exc_info=True)
                return Response({'error': f"Error al modificar el archivo '{file_name}': {e}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # 5. Llamar al endpoint de importación
        if not archivos_procesados_para_importar:
            return Response({'message': 'No hay archivos para importar después del procesamiento.'}, status=status.HTTP_200_OK)

        try:
            # NOTA: Asegúrate de que los servicios se puedan comunicar.
            # Si se ejecutan en el mismo servidor/contenedor, 'localhost' funciona.
            # En Docker Compose, usarías el nombre del servicio, ej: 'http://gestion-asignacion:8000/...'
            import_url = request.build_absolute_uri('http://gestion-asignacion:8000/inventario/api/importar-perfiles/')
            
            response = requests.post(import_url, files=archivos_procesados_para_importar)
            response.raise_for_status()

            # Limpiar la caché después de una importación exitosa
            cache.delete(cache_key)

            return Response(response.json(), status=response.status_code)

        except requests.RequestException as e:
            error_data = {'error': f'Error al llamar al servicio de importación: {e}'}
            if e.response is not None:
                try:
                    error_data['detalle_servicio'] = e.response.json()
                except ValueError:
                    error_data['detalle_servicio'] = e.response.text
            return Response(error_data, status=status.HTTP_503_SERVICE_UNAVAILABLE)





class PrevisualizarIncompletosAPIView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, format=None):
        excel_files = request.FILES.getlist('files')
        if not excel_files:
            return Response({'error': 'No se encontraron archivos en la solicitud.'}, status=status.HTTP_400_BAD_REQUEST)

        filas_incompletas = []
        errores_procesamiento = []

        # Columnas que deben estar vacías para que una fila se considere "incompleta"
        # (además de que 'CODIGO' debe estar lleno)
        columnas_a_verificar = [
            'CREDITOS_APROBADOS',
            'PROMEDIO_CARRERA',
            'APROBADAS',
            'PERIODOS_MATRICULADOS'
        ]

        # FUNCIÓN AUXILIAR MEJORADA: Detectar si una fila está completamente vacía
        def es_fila_vacia(row_dict):
            """
            Retorna True si todos los valores VÁLIDOS de la fila están vacíos.
            Ignora las claves None y solo evalúa las columnas con nombres válidos.
            """
            if not row_dict:
                return True
            
            # Filtrar solo las claves que son strings (columnas válidas)
            valores_validos = [
                val for key, val in row_dict.items() 
                if key is not None and isinstance(key, str)
            ]
            
            # Si no hay valores válidos, la fila está vacía
            if not valores_validos:
                return True
            
            # Verificar si todos los valores válidos están vacíos
            return all(
                val is None or str(val).strip() == '' 
                for val in valores_validos
            )

        for excel_file in excel_files:
            try:
                filename = excel_file.name.lower()
                if filename.endswith('.xlsx'):
                    file_format = 'xlsx'
                elif filename.endswith('.xls'):
                    file_format = 'xls'
                else:
                    errores_procesamiento.append(f"Archivo '{excel_file.name}' ignorado: formato no soportado.")
                    continue

                excel_file.seek(0)
                
                dataset = Dataset()
                dataset.load(excel_file.read(), format=file_format)

                total_filas = len(dataset)
                logger.info(f"=== Previsualizar - Archivo '{excel_file.name}' cargado. Total filas: {total_filas} ===")

                filas_procesadas = 0
                filas_vacias_saltadas = 0
                filas_sin_codigo = 0
                filas_completas = 0
                filas_incompletas_encontradas = 0

                for i, row in enumerate(dataset.dict):
                    filas_procesadas += 1
                    
                    # Log cada 100 filas para monitorear progreso
                    if filas_procesadas % 100 == 0:
                        logger.info(f"Progreso: {filas_procesadas}/{total_filas} filas procesadas...")
                    
                    # SOLUCIÓN: Saltar filas completamente vacías PRIMERO
                    if es_fila_vacia(row):
                        filas_vacias_saltadas += 1
                        continue

                    codigo_valor = row.get('CODIGO')

                    # Validar que el código sea un número válido
                    try:
                        float(codigo_valor)
                    except (ValueError, TypeError):
                        filas_sin_codigo += 1
                        continue
                    else:
                        def is_cell_empty(value):
                            """
                            Función robusta para determinar si una celda está vacía.
                            Devuelve True para: None, '', '   '.
                            Devuelve False para: 'texto', 0, 123, '0'.
                            """
                            return value is None or not str(value).strip()

                        otras_columnas_vacias = any(is_cell_empty(row.get(col)) for col in columnas_a_verificar)

                        if otras_columnas_vacias:
                            # Limpiamos el código para asegurar que sea un entero
                            str_value = str(codigo_valor)
                            if str_value.endswith('.0'):
                                str_value = str_value[:-2]
                            
                            filas_incompletas_encontradas += 1
                            filas_incompletas.append({
                                'codigo': int(str_value),
                                'fila': i + 2,
                                'archivo': excel_file.name
                            })
                        else:
                            filas_completas += 1
                
                # Resumen del archivo procesado
                logger.info(f"=== Resumen de '{excel_file.name}' ===")
                logger.info(f"  Total filas: {total_filas}")
                logger.info(f"  Filas vacías saltadas: {filas_vacias_saltadas}")
                logger.info(f"  Filas sin código válido: {filas_sin_codigo}")
                logger.info(f"  Filas completas: {filas_completas}")
                logger.info(f"  Filas incompletas: {filas_incompletas_encontradas}")
                            
            except Exception as e:
                msg = f"Error crítico al procesar el archivo '{excel_file.name}': {e}"
                logger.error(msg, exc_info=True)
                errores_procesamiento.append(msg)

        logger.info(f"=== PROCESO COMPLETO ===")
        logger.info(f"Total de filas incompletas encontradas en todos los archivos: {len(filas_incompletas)}")
        
        return Response({'filas_incompletas': filas_incompletas, 'advertencias': errores_procesamiento}, status=status.HTTP_200_OK)




class ImportarProductosAPIView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, format=None):
        """
        Recibe un archivo Excel (.xlsx o .xls), valida su contenido y lo importa/actualiza PerfilAcademico.
        """
        excel_files = request.FILES.getlist('files')

        if not excel_files:
            return Response({'error': 'No se encontraron archivos en la solicitud.'}, 
                            status=status.HTTP_400_BAD_REQUEST)
        
        # ========== INICIO: OBTENER PERIODO ACTIVO ==========
        try:
            periodo_response = requests.get('http://gestion-asignacion:8000/api/asignacion/procesos/periodo-activo/')
            if periodo_response.status_code == 204:
                return Response({'error': 'No se encontró un proceso de asignación ACTIVO.'}, 
                              status=status.HTTP_404_NOT_FOUND)
            periodo_response.raise_for_status()
            periodo_data = periodo_response.json()
            anio_activo = periodo_data['pa_anio']
            semestre_activo = periodo_data['pa_num_semestre']
            logger.info(f"Periodo activo obtenido: {anio_activo}-{semestre_activo}")
        except requests.RequestException as e:
            return Response({'error': f'No se pudo comunicar con el servicio de asignación: {e}'}, 
                          status=status.HTTP_503_SERVICE_UNAVAILABLE)
        # ========== FIN: OBTENER PERIODO ACTIVO ==========
        
        # FUNCIÓN AUXILIAR: Detectar si una fila está completamente vacía
        def es_fila_vacia(row_dict):
            """
            Retorna True si todos los valores VÁLIDOS de la fila están vacíos.
            """
            if not row_dict:
                return True
            
            # Filtrar solo las claves que son strings (columnas válidas)
            valores_validos = [
                val for key, val in row_dict.items() 
                if key is not None and isinstance(key, str)
            ]
            
            if not valores_validos:
                return True
            
            return all(
                val is None or str(val).strip() == '' 
                for val in valores_validos
            )
        
        # FUNCIÓN AUXILIAR: Validar que una fila tiene un código válido
        def tiene_codigo_valido(row_dict):
            """
            Retorna True si la fila tiene un código de estudiante válido.
            """
            codigo = row_dict.get('CODIGO')
            if codigo is None:
                return False
            try:
                float(codigo)
                return True
            except (ValueError, TypeError):
                return False
        
        # Creamos un Dataset agregado para combinar los datos de todos los archivos.
        dataset = Dataset()
        first_file = True
        filas_filtradas = 0
        filas_totales = 0

        for excel_file in excel_files:
            try:
                excel_file.seek(0)
                file_content = excel_file.read()
                current_dataset = Dataset()
                
                error_xlsx = None
                error_xls = None

                try:
                    current_dataset.load(file_content, format='xlsx')
                    logger.info(f"Archivo '{excel_file.name}' leído exitosamente como .xlsx")
                except Exception as e_xlsx:
                    error_xlsx = e_xlsx
                    try:
                        logger.warning(f"Fallo al leer '{excel_file.name}' como .xlsx, reintentando como .xls...")
                        current_dataset.load(file_content, format='xls')
                        logger.info(f"Archivo '{excel_file.name}' leído exitosamente como .xls")
                    except Exception as e_xls:
                        error_xls = e_xls
                        raise ValueError(f"No se pudo leer como .xlsx ({repr(error_xlsx)}) ni como .xls ({repr(error_xls)}). El archivo podría estar corrupto o en un formato no soportado.")

                if first_file:
                    dataset.headers = current_dataset.headers
                    first_file = False
                
                # FILTRAR FILAS VACÍAS Y SIN CÓDIGO ANTES DE AGREGAR AL DATASET
                for row_dict in current_dataset.dict:
                    filas_totales += 1
                    
                    # Saltar filas vacías
                    if es_fila_vacia(row_dict):
                        filas_filtradas += 1
                        continue
                    
                    # Saltar filas sin código válido
                    if not tiene_codigo_valido(row_dict):
                        filas_filtradas += 1
                        continue
                    
                    # Solo agregar filas válidas
                    new_row = tuple(row_dict.get(header) for header in dataset.headers)
                    dataset.append(new_row)

            except Exception as e:
                logger.error(f"Error crítico al leer el archivo '{excel_file.name}' para importación.", exc_info=True)
                return Response({'error': f"Error al procesar el archivo '{excel_file.name}': {repr(e)}"}, 
                                status=status.HTTP_400_BAD_REQUEST)

        logger.info(f"Dataset preparado: {len(dataset)} filas válidas de {filas_totales} totales ({filas_filtradas} filtradas)")

        # 3. Prueba de importación (Dry Run)
        resource = PerfilAcademicoResource()
        
        # ========== PASAR PERIODO AL RESOURCE ==========
        kwargs = {
            'anio_activo': anio_activo,
            'semestre_activo': semestre_activo
        }
        
        try:
            result = resource.import_data(
                dataset, 
                dry_run=True,           
                raise_errors=False,     
                use_transactions=True,
                **kwargs
            ) 
        except Exception as e:
            return Response({'error': f'Error de validación general: {e}', 'trace': traceback.format_exc()}, 
                            status=status.HTTP_400_BAD_REQUEST)

        if result.has_errors() or result.has_validation_errors():
            error_details = []
            MAX_ERRORES = 50
            
            error_count = 0
            for row in result.row_errors():
                if error_count >= MAX_ERRORES:
                    break
                    
                fila_excel = row[0] + 2
                errores_de_fila = []
                for error in row[1]:
                    if hasattr(error.error, 'items'):
                        for field, messages in error.error.items():
                            errores_de_fila.append(f"Campo '{field}': {messages[0]}")
                    elif hasattr(error.error, 'message'):
                        errores_de_fila.append(error.error.message)
                    else:
                        errores_de_fila.append(str(error.error))
                error_details.append({
                    'fila': fila_excel, 
                    'errores': errores_de_fila
                })
                error_count += 1
            
            for error in result.base_errors:
                if error_count >= MAX_ERRORES:
                    break
                error_details.append({'fila': 'General', 'errores': [str(error.error)]})
                error_count += 1

            total_errores = len(list(result.row_errors())) + len(result.base_errors)
            errores_omitidos = max(0, total_errores - MAX_ERRORES)
            
            response_data = {
                'message': 'Importación fallida. Se encontraron errores de datos/mapeo.',
                'total_errores': total_errores,
                'errores_mostrados': len(error_details),
                'errores': error_details
            }
            
            if errores_omitidos > 0:
                response_data['advertencia'] = f'Se omitieron {errores_omitidos} errores adicionales. Solo se muestran los primeros {MAX_ERRORES} errores.'
            
            logger.warning(f"Importación fallida con {total_errores} errores. Mostrando primeros {len(error_details)}.")
            
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

        # 4. Importación final (Guardar en la BD)
        try:
            result_final = resource.import_data(
                dataset, 
                dry_run=False,          
                raise_errors=True,      
                use_transactions=True,
                **kwargs
            ) 
        except Exception as e:
            return Response({'error': f'Error al guardar en la base de datos: {e}', 'trace': traceback.format_exc()}, 
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # 5. Respuesta exitosa
        return Response({
            'message': '¡Perfiles académicos importados y mapeados exitosamente!',
            'periodo': f'{anio_activo}-{semestre_activo}',
            'resumen': {
                'total_registros': len(dataset),
                'filas_filtradas': filas_filtradas,
                'creados': result_final.totals.get('new', 0),
                'actualizados': result_final.totals.get('update', 0)
            }
        }, status=status.HTTP_200_OK)